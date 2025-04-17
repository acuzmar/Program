import os
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import requests
import urllib3
import json
import logging
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Para manejar la imagen de fondo
from PIL import Image, ImageTk
from PIL.Image import Resampling

###################################################################
# 1. CONFIGURACIONES INICIALES
###################################################################
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)

F5_IP_GLOBAL = ""
F5_USER_GLOBAL = ""
F5_PASS_GLOBAL = ""

CF_EMAIL_GLOBAL = ""
CF_TOKEN_GLOBAL = ""

ENV_USER = os.getenv("F5_USER", "")
ENV_PASS = os.getenv("F5_PASS", "")

ENV_CF_EMAIL = os.getenv("CF_EMAIL", "")
ENV_CF_TOKEN = os.getenv("CF_API_TOKEN", "")

CA_FILE_PATH = None

# Estructura de datos global:
# {
#   "f5": { "basic": {...}, "advanced": {...}, "monitors": {...} },
#   "cloudflare": { "zones": {...}, "dns_records": {...}, "extra": {...} }
# }
f5_data_store = {
    "f5": {
        "basic": {},
        "advanced": {},
        "monitors": {}
    },
    "cloudflare": {}
}

###################################################################
# 2. DEFINIR ENDPOINTS F5
###################################################################
ENDPOINTS_BASIC = {
    "nodes":       "/mgmt/tm/ltm/node",
    "virtuals":    "/mgmt/tm/ltm/virtual",
    "pools":       "/mgmt/tm/ltm/pool",
    "certs":       "/mgmt/tm/sys/crypto/cert"
}

ENDPOINTS_ADVANCED = {
    "stats": {
        "virtual_stats": "/mgmt/tm/ltm/virtual/stats",
        "node_stats":    "/mgmt/tm/ltm/node/stats",
        "pool_stats":    "/mgmt/tm/ltm/pool/stats",
    },
    "system": {
        "cm_device":        "/mgmt/tm/cm/device",
        "failoverStatus":   "/mgmt/tm/cm/failoverStatus",
        "license":          "/mgmt/tm/sys/license",
        "version":          "/mgmt/tm/sys/version",
        "host_info":        "/mgmt/tm/sys/host-info",
        "available_memory": "/mgmt/tm/sys/available-memory",
    },
    "networking": {
        "net_vlan": "/mgmt/tm/net/vlan",
        "net_self": "/mgmt/tm/net/self",
        "net_route": "/mgmt/tm/net/route",
        "net_arp":   "/mgmt/tm/net/arp",
        "ltm_nat":   "/mgmt/tm/ltm/nat",
        "ltm_snat":  "/mgmt/tm/ltm/snat",
    },
    "profiles": {
        "client_ssl": "/mgmt/tm/ltm/profile/client-ssl",
        "server_ssl": "/mgmt/tm/ltm/profile/server-ssl",
        "profile_http": "/mgmt/tm/ltm/profile/http",
        "profile_tcp":  "/mgmt/tm/ltm/profile/tcp",
        "ltm_rule":     "/mgmt/tm/ltm/rule",
        "ltm_policy":   "/mgmt/tm/ltm/policy"
    },
    "modules": {
        "apm_profile_access": "/mgmt/tm/apm/profile/access",
        "apm_session":        "/mgmt/tm/apm/session/",
        "asm_policies":       "/mgmt/tm/asm/policies",
        "asm_logs":           "/mgmt/tm/asm/logs",
        "gtm_wideip":         "/mgmt/tm/gtm/wideip",
        "gtm_pool_a_members": "/mgmt/tm/gtm/pool/a/members",
        "gtm_wideip_a_stats": "/mgmt/tm/gtm/wideip/a/stats",
    },
    "logs_config": {
        "sys_log_config": "/mgmt/tm/sys/log-config/destination"
    },
    "ha_cluster": {
        "device_group": "/mgmt/tm/cm/device-group",
        "sync_status":  "/mgmt/tm/cm/sync-status",
        "trust_domain": "/mgmt/tm/cm/trust-domain"
    }
}

# NUEVO: ENDPOINTS PARA MONITORES, iRules y Persistencia
ENDPOINTS_MONITORS = {
    "monitors":    "/mgmt/tm/ltm/monitor",
    "irules":      "/mgmt/tm/ltm/irule",
    "persistence": "/mgmt/tm/ltm/persistence"
}

###################################################################
# 2B. DEFINIR ENDPOINTS CLOUDFLARE
###################################################################
# Endpoints básicos y extendidos para Cloudflare:
CLOUDFLARE_ENDPOINTS = {
    "list_zones":         "/client/v4/zones",
    "dns_records":        "/client/v4/zones/{zone_id}/dns_records",
    "analytics_dashboard": "/client/v4/zones/{zone_id}/analytics/dashboard",
    "analytics_colos":     "/client/v4/zones/{zone_id}/analytics/colos",
    "firewall_settings":   "/client/v4/zones/{zone_id}/firewall/settings",
    "firewall_rules":      "/client/v4/zones/{zone_id}/firewall/rules",
    "firewall_events":     "/client/v4/zones/{zone_id}/firewall/events",
    "ssl_settings":        "/client/v4/zones/{zone_id}/ssl",
    "zone_settings":       "/client/v4/zones/{zone_id}/settings",
    "rate_limits":         "/client/v4/zones/{zone_id}/rate_limits",
    "dnssec":              "/client/v4/zones/{zone_id}/dnssec",
    "load_balancers":      "/client/v4/zones/{zone_id}/load_balancers"
}

###################################################################
# 3. FUNCIONES DE CONEXIÓN (F5 / CF)
###################################################################
def get_f5_data(base_url, endpoint, auth, ssl_verify):
    if ssl_verify is False:
        urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    try:
        full_url = base_url + endpoint
        resp = requests.get(full_url, auth=auth, verify=ssl_verify)
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as e:
        logging.warning(f"Error al conectar con {full_url}: {e}")
        return None

def collect_basic_data(ip, user, password, ssl_verify, partition):
    base_url = f"https://{ip}"
    auth = (user, password)
    partition_filter = f"?$filter=partition eq {partition.strip()}" if partition.strip() else ""
    data_basic = {}
    for key, ep in ENDPOINTS_BASIC.items():
        endpoint = ep + partition_filter if key in ["nodes", "virtuals", "pools"] and partition_filter else ep
        data_basic[key] = get_f5_data(base_url, endpoint, auth, ssl_verify)
    return data_basic

def collect_advanced_data(ip, user, password, ssl_verify):
    base_url = f"https://{ip}"
    auth = (user, password)
    data_advanced = {}
    for category, cat_endpoints in ENDPOINTS_ADVANCED.items():
        cat_dict = {}
        for key, ep in cat_endpoints.items():
            cat_dict[key] = get_f5_data(base_url, ep, auth, ssl_verify)
        data_advanced[category] = cat_dict
    return data_advanced

def collect_monitors_data(ip, user, password, ssl_verify):
    base_url = f"https://{ip}"
    auth = (user, password)
    data_monitors = {}
    for key, ep in ENDPOINTS_MONITORS.items():
        data_monitors[key] = get_f5_data(base_url, ep, auth, ssl_verify)
    return data_monitors

def get_cloudflare_data(endpoint, email, token, params=None):
    base_url = "https://api.cloudflare.com"
    url = base_url + endpoint
    headers = {
        "X-Auth-Email": email,
        "X-Auth-Key": token,
        "Content-Type": "application/json"
    }
    try:
        resp = requests.get(url, headers=headers, params=params)
        resp.raise_for_status()
        return resp.json()
    except requests.RequestException as e:
        logging.warning(f"[Cloudflare] Error al conectar con {url}: {e}")
        return None

def collect_cf_zones(email, token):
    endpoint = CLOUDFLARE_ENDPOINTS["list_zones"]
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_dns_records(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["dns_records"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_analytics_dashboard(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["analytics_dashboard"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_analytics_colos(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["analytics_colos"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_firewall_settings(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["firewall_settings"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_firewall_rules(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["firewall_rules"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_firewall_events(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["firewall_events"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_ssl_settings(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["ssl_settings"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_zone_settings(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["zone_settings"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_rate_limits(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["rate_limits"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_dnssec(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["dnssec"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

def collect_cf_load_balancers(email, token, zone_id):
    endpoint = CLOUDFLARE_ENDPOINTS["load_balancers"].replace("{zone_id}", zone_id)
    return get_cloudflare_data(endpoint, email, token)

###################################################################
# 4. FUNCIONES DE GUI - RECOLECCIÓN
###################################################################
def run_basic_collection():
    text_area.delete("1.0", tk.END)
    ip = F5_IP_GLOBAL
    user = F5_USER_GLOBAL or ENV_USER
    password = F5_PASS_GLOBAL or ENV_PASS
    if not ip or not user or not password:
        messagebox.showwarning("Faltan credenciales / IP", "Ve a la pestaña LOGIN (F5).")
        return
    partition = entry_part_b.get().strip()
    ssl_verify = CA_FILE_PATH if (ssl_verify_var.get() and CA_FILE_PATH) else (True if ssl_verify_var.get() else False)
    text_area.insert(tk.END, f"[BASIC-F5] Conectando a {ip}\n")
    data_b = collect_basic_data(ip, user, password, ssl_verify, partition)
    f5_data_store["f5"]["basic"] = data_b
    nodes_items = data_b.get("nodes", {}).get("items", [])
    vs_items    = data_b.get("virtuals", {}).get("items", [])
    pools_items = data_b.get("pools", {}).get("items", [])
    certs_items = data_b.get("certs", {}).get("items", [])
    text_area.insert(tk.END, f"IP: {ip}\n")
    text_area.insert(tk.END, f"  Nodos: {len(nodes_items)}\n")
    text_area.insert(tk.END, f"  VS: {len(vs_items)}\n")
    text_area.insert(tk.END, f"  Pools: {len(pools_items)}\n")
    text_area.insert(tk.END, f"  Certs: {len(certs_items)}\n\n")
    text_area.insert(tk.END, "¡Recolección BÁSICA (F5) finalizada!\n")

def run_advanced_collection():
    text_area.delete("1.0", tk.END)
    ip = F5_IP_GLOBAL
    user = F5_USER_GLOBAL or ENV_USER
    password = F5_PASS_GLOBAL or ENV_PASS
    if not ip or not user or not password:
        messagebox.showwarning("Faltan credenciales / IP", "Ve a la pestaña LOGIN (F5).")
        return
    ssl_verify = CA_FILE_PATH if (ssl_verify_var.get() and CA_FILE_PATH) else (True if ssl_verify_var.get() else False)
    text_area.insert(tk.END, f"[ADVANCED-F5] Conectando a {ip}\n")
    data_adv = collect_advanced_data(ip, user, password, ssl_verify)
    f5_data_store["f5"]["advanced"] = data_adv
    text_area.insert(tk.END, f"IP: {ip} - Datos avanzados obtenidos\n\n")
    text_area.insert(tk.END, "¡Recolección AVANZADA (F5) finalizada!\n")

def run_monitors_collection():
    text_area.delete("1.0", tk.END)
    ip = F5_IP_GLOBAL
    user = F5_USER_GLOBAL or ENV_USER
    password = F5_PASS_GLOBAL or ENV_PASS
    if not ip or not user or not password:
        messagebox.showwarning("Faltan credenciales / IP", "Ve a la pestaña LOGIN (F5).")
        return
    ssl_verify = CA_FILE_PATH if (ssl_verify_var.get() and CA_FILE_PATH) else (True if ssl_verify_var.get() else False)
    text_area.insert(tk.END, f"[MONITORES-F5] Conectando a {ip}\n")
    monitors_data = collect_monitors_data(ip, user, password, ssl_verify)
    f5_data_store["f5"]["monitors"] = monitors_data
    for key, data in monitors_data.items():
        text_area.insert(tk.END, f"--- {key.upper()} ---\n")
        text_area.insert(tk.END, json.dumps(data, indent=2) + "\n\n")
    text_area.insert(tk.END, "¡Recolección de MONITORES finalizada!\n")

def run_cloudflare_collection():
    text_area.delete("1.0", tk.END)
    email = CF_EMAIL_GLOBAL or ENV_CF_EMAIL
    token = CF_TOKEN_GLOBAL or ENV_CF_TOKEN
    if not email or not token:
        messagebox.showwarning("Cloudflare", "Faltan Email / Token. Ve a la pestaña LOGIN CF.")
        return
    text_area.insert(tk.END, "[Cloudflare] Consultando Zonas...\n")
    zones_data = collect_cf_zones(email, token)
    f5_data_store["cloudflare"]["zones"] = zones_data
    if zones_data and zones_data.get("success", False):
        zones_result = zones_data.get("result", [])
        text_area.insert(tk.END, f"Zonas encontradas: {len(zones_result)}\n")
        for z in zones_result:
            z_name = z.get("name", "")
            z_id   = z.get("id", "")
            text_area.insert(tk.END, f" - Zone: {z_name} (ID: {z_id})\n")
        if zones_result:
            first_zone = zones_result[0]
            zone_id = first_zone.get("id", "")
            text_area.insert(tk.END, "\nConsultando DNS Records...\n")
            dns_data = collect_cf_dns_records(email, token, zone_id)
            f5_data_store["cloudflare"]["dns_records"] = dns_data
            if dns_data and dns_data.get("success", False):
                dns_result = dns_data.get("result", [])
                text_area.insert(tk.END, f"DNS Records encontrados: {len(dns_result)}\n")
                for d in dns_result:
                    rec_name = d.get("name", "")
                    rec_type = d.get("type", "")
                    rec_content = d.get("content", "")
                    text_area.insert(tk.END, f" - {rec_name} ({rec_type}): {rec_content}\n")
            else:
                text_area.insert(tk.END, "No se pudo obtener DNS Records.\n")
    else:
        text_area.insert(tk.END, "No se pudo obtener la lista de Zonas.\n")
    text_area.insert(tk.END, "\n¡Consulta Cloudflare finalizada!\n")

def run_cf_extra_collection():
    text_area.delete("1.0", tk.END)
    email = CF_EMAIL_GLOBAL or ENV_CF_EMAIL
    token = CF_TOKEN_GLOBAL or ENV_CF_TOKEN
    if not email or not token:
        messagebox.showwarning("Cloudflare", "Faltan Email / Token. Ve a la pestaña LOGIN CF.")
        return
    zones_data = f5_data_store["cloudflare"].get("zones", {})
    if zones_data and zones_data.get("success", False):
        zones_result = zones_data.get("result", [])
        if zones_result:
            first_zone = zones_result[0]
            zone_id = first_zone.get("id", "")
            text_area.insert(tk.END, f"Usando la zona: {first_zone.get('name', '')}\n\n")
            extra = {}
            extra["analytics_dashboard"] = collect_cf_analytics_dashboard(email, token, zone_id)
            extra["analytics_colos"]     = collect_cf_analytics_colos(email, token, zone_id)
            extra["firewall_settings"]   = collect_cf_firewall_settings(email, token, zone_id)
            extra["firewall_rules"]      = collect_cf_firewall_rules(email, token, zone_id)
            extra["firewall_events"]     = collect_cf_firewall_events(email, token, zone_id)
            extra["ssl_settings"]        = collect_cf_ssl_settings(email, token, zone_id)
            extra["zone_settings"]       = collect_cf_zone_settings(email, token, zone_id)
            extra["rate_limits"]         = collect_cf_rate_limits(email, token, zone_id)
            extra["dnssec"]              = collect_cf_dnssec(email, token, zone_id)
            extra["load_balancers"]      = collect_cf_load_balancers(email, token, zone_id)
            f5_data_store["cloudflare"]["extra"] = extra
            for key, data in extra.items():
                text_area.insert(tk.END, f"--- {key.upper()} ---\n")
                # Verifica si data tiene clave "result" y es una lista
                if isinstance(data, dict) and "result" in data:
                    items = data.get("result", [])
                    if isinstance(items, list) and len(items) > 0:
                        text_area.insert(tk.END, json.dumps(items, indent=2) + "\n\n")
                    else:
                        text_area.insert(tk.END, "Sin datos\n\n")
                else:
                    text_area.insert(tk.END, json.dumps(data, indent=2) + "\n\n")
            text_area.insert(tk.END, "¡Recolección de CF-EXTRA finalizada!\n")
        else:
            text_area.insert(tk.END, "No hay zonas disponibles para consultar datos extra.\n")
    else:
        text_area.insert(tk.END, "No se pudo obtener la lista de Zonas.\n")

###################################################################
# 4C. LOGIN F5 / CF
###################################################################
def save_login_f5():
    global F5_IP_GLOBAL, F5_USER_GLOBAL, F5_PASS_GLOBAL
    ip = entry_login_ip.get().strip()
    user = entry_login_user.get().strip()
    password = entry_login_pass.get().strip()
    if not ip or not user or not password:
        messagebox.showwarning("Login F5", "Faltan IP, usuario o contraseña.")
        return
    F5_IP_GLOBAL = ip
    F5_USER_GLOBAL = user
    F5_PASS_GLOBAL = password
    messagebox.showinfo("Login F5", f"¡Login guardado!\nIP: {F5_IP_GLOBAL}\nUsuario: {F5_USER_GLOBAL}")

def save_login_cf():
    global CF_EMAIL_GLOBAL, CF_TOKEN_GLOBAL
    email = entry_cf_email.get().strip()
    token = entry_cf_token.get().strip()
    if not email or not token:
        messagebox.showwarning("Login CF", "Faltan Email / Token.")
        return
    CF_EMAIL_GLOBAL = email
    CF_TOKEN_GLOBAL = token
    messagebox.showinfo("Login Cloudflare", f"¡Login guardado!\nEmail: {CF_EMAIL_GLOBAL}")

###################################################################
# 5. EXPORTAR A EXCEL (F5 + Cloudflare + CF-EXTRA) con verificación final
###################################################################
def export_to_excel():
    has_f5_basic = bool(f5_data_store["f5"].get("basic"))
    has_f5_adv   = bool(f5_data_store["f5"].get("advanced"))
    has_f5_mon   = bool(f5_data_store["f5"].get("monitors"))
    has_f5 = has_f5_basic or has_f5_adv or has_f5_mon
    cf_zones = bool(f5_data_store["cloudflare"].get("zones"))
    cf_dns   = bool(f5_data_store["cloudflare"].get("dns_records"))
    has_cf = cf_zones or cf_dns
    cf_extra = bool(f5_data_store["cloudflare"].get("extra"))
    logging.info("DEBUG al Exportar: f5_data_store => \n" + json.dumps(f5_data_store, indent=2))
    if not has_f5 and not has_cf and not cf_extra:
        messagebox.showinfo("Sin datos", "No hay datos que exportar. Primero ejecuta una recolección.")
        return
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Archivo Excel", "*.xlsx"), ("Todos los archivos", "*.*")]
    )
    if not file_path:
        return
    wb = Workbook()
    default_sheet = wb["Sheet"]
    wb.remove(default_sheet)
    
    # ---------- EXPORTACIÓN F5 ----------
    if has_f5:
        f5_info = f5_data_store["f5"]
        basic_info = f5_info.get("basic", {})
        ws_nodes = wb.create_sheet("F5-Nodos")
        ws_nodes.append(["Name", "Address", "Session", "Monitor"])
        items_nodes = basic_info.get("nodes", {}).get("items", [])
        if isinstance(items_nodes, list) and len(items_nodes) > 0:
            for node in items_nodes:
                name = node.get("name", "")
                address = node.get("address", "")
                session = node.get("session", "")
                monitor = node.get("state", "")
                ws_nodes.append([name, address, session, monitor])
        else:
            ws_nodes.append(["Sin datos"])
        auto_adjust_columns(ws_nodes)

        ws_vs = wb.create_sheet("F5-VS")
        ws_vs.append(["Name", "Destination", "Pool", "Enabled"])
        items_vs = basic_info.get("virtuals", {}).get("items", [])
        if isinstance(items_vs, list) and len(items_vs) > 0:
            for vs in items_vs:
                vs_name = vs.get("name", "")
                dest = vs.get("destination", "")
                ip_port = dest.split("/")[-1] if "/" in dest else dest
                pool = vs.get("pool", "N/A").split("/")[-1] if "pool" in vs else "N/A"
                enabled_state = "Enabled" if vs.get("enabled", True) else "Disabled"
                ws_vs.append([vs_name, ip_port, pool, enabled_state])
        else:
            ws_vs.append(["Sin datos"])
        auto_adjust_columns(ws_vs)

        ws_pools = wb.create_sheet("F5-Pools")
        ws_pools.append(["Name", "LB Mode", "Monitor"])
        items_pools = basic_info.get("pools", {}).get("items", [])
        if isinstance(items_pools, list) and len(items_pools) > 0:
            for pool in items_pools:
                p_name = pool.get("name", "")
                lb_mode = pool.get("loadBalancingMode", "")
                monitor = pool.get("monitor", "")
                ws_pools.append([p_name, lb_mode, monitor])
        else:
            ws_pools.append(["Sin datos"])
        auto_adjust_columns(ws_pools)

        ws_certs = wb.create_sheet("F5-Certs")
        ws_certs.append(["Name", "Partition"])
        items_certs = basic_info.get("certs", {}).get("items", [])
        if isinstance(items_certs, list) and len(items_certs) > 0:
            for cert in items_certs:
                c_name = cert.get("name", "")
                partition = cert.get("partition", "")
                ws_certs.append([c_name, partition])
        else:
            ws_certs.append(["Sin datos"])
        auto_adjust_columns(ws_certs)

        adv_info = f5_info.get("advanced", {})
        for cat_name, cat_data in adv_info.items():
            for sub_key, sub_json in cat_data.items():
                sheet_name = f"F5-{cat_name}-{sub_key}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                ws_adv = wb.create_sheet(sheet_name)
                if sub_json is None:
                    ws_adv.append(["No hay datos (None)."])
                else:
                    if "stats" in sub_key.lower():
                        if isinstance(sub_json, dict) and "entries" in sub_json:
                            ws_adv.append(["Key", "Availability", "Enabled", "Extra Info"])
                            for k_obj, v_obj in sub_json["entries"].items():
                                nested = v_obj.get("nestedStats", {}).get("entries", {})
                                availability = nested.get("status.availabilityState", {}).get("description", "")
                                enabled = nested.get("status.enabledState", {}).get("description", "")
                                extra = json.dumps(nested, indent=1)
                                ws_adv.append([k_obj, availability, enabled, extra])
                        else:
                            ws_adv.append(["Respuesta no parseable para stats."])
                            ws_adv.append([json.dumps(sub_json, indent=2)])
                    elif isinstance(sub_json, dict) and "items" in sub_json:
                        items = sub_json.get("items")
                        if isinstance(items, list) and len(items) > 0:
                            headers = list(items[0].keys())
                            ws_adv.append(headers)
                            for it in items:
                                row_data = [str(it.get(h, "")) for h in headers]
                                ws_adv.append(row_data)
                        else:
                            ws_adv.append(["Sin items en este endpoint."])
                    elif isinstance(sub_json, dict) and "entries" in sub_json:
                        entries = sub_json.get("entries")
                        ws_adv.append(["Key", "Value"])
                        for k, v in entries.items():
                            ws_adv.append([k, json.dumps(v, indent=1)])
                    else:
                        ws_adv.append(["JSON Bruto"])
                        ws_adv.append([json.dumps(sub_json, indent=2)])
                auto_adjust_columns(ws_adv)

        monitors_data = f5_info.get("monitors", {})
        if monitors_data:
            for key, data in monitors_data.items():
                sheet_name = f"F5-Monitors-{key}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                ws_mon = wb.create_sheet(sheet_name)
                if isinstance(data, dict) and "items" in data:
                    items = data.get("items")
                    if isinstance(items, list) and len(items) > 0:
                        headers = list(items[0].keys())
                        ws_mon.append(headers)
                        for it in items:
                            row_data = [str(it.get(h, "")) for h in headers]
                            ws_mon.append(row_data)
                    else:
                        ws_mon.append(["Sin items en este endpoint."])
                else:
                    ws_mon.append([json.dumps(data, indent=2)])
                auto_adjust_columns(ws_mon)

    # ---------- EXPORTACIÓN CLOUDFLARE ----------
    if has_cf:
        cf_data = f5_data_store["cloudflare"]
        zones_data = cf_data.get("zones", {})
        ws_zones = wb.create_sheet("CF-Zones")
        ws_zones.append(["ID", "Name", "Status", "DevMode", "Plan"])
        if zones_data and zones_data.get("success", False):
            for z in zones_data.get("result", []):
                z_id = z.get("id", "")
                z_name = z.get("name", "")
                z_status = z.get("status", "")
                z_dev = z.get("developmentMode", "")
                z_plan = z.get("plan", {}).get("name", "")
                ws_zones.append([z_id, z_name, z_status, z_dev, z_plan])
        else:
            ws_zones.append(["No se pudo obtener Zonas."])
        auto_adjust_columns(ws_zones)

        dns_data = cf_data.get("dns_records", {})
        ws_dns = wb.create_sheet("CF-DNSRecords")
        ws_dns.append(["ID", "Name", "Type", "Content", "TTL"])
        if dns_data and dns_data.get("success", False):
            for rec in dns_data.get("result", []):
                r_id = rec.get("id", "")
                r_name = rec.get("name", "")
                r_type = rec.get("type", "")
                r_content = rec.get("content", "")
                r_ttl = rec.get("ttl", "")
                ws_dns.append([r_id, r_name, r_type, r_content, r_ttl])
        else:
            ws_dns.append(["No se pudo obtener DNS Records."])
        auto_adjust_columns(ws_dns)

        cf_extra = f5_data_store["cloudflare"].get("extra", {})
        if cf_extra:
            for key, data in cf_extra.items():
                sheet_name = f"CF-Extra-{key}"
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                ws_extra = wb.create_sheet(sheet_name)
                if isinstance(data, dict) and "result" in data:
                    items = data.get("result", [])
                    if isinstance(items, list) and len(items) > 0:
                        headers = list(items[0].keys())
                        ws_extra.append(headers)
                        for it in items:
                            row_data = [str(it.get(h, "")) for h in headers]
                            ws_extra.append(row_data)
                    else:
                        ws_extra.append(["Sin datos"])
                else:
                    ws_extra.append([json.dumps(data, indent=2)])
                auto_adjust_columns(ws_extra)

    # Si por alguna razón no se creó ninguna hoja, creamos "EmptySheet"
    if len(wb.sheetnames) == 0:
        wb.create_sheet("EmptySheet")
    try:
        wb.save(file_path)
        messagebox.showinfo("Exportar a Excel", f"Datos exportados correctamente en:\n{file_path}")
    except Exception as e:
        messagebox.showerror("Error al guardar Excel", str(e))

###################################################################
# 6. FUNCIÓN AUXILIAR: AUTO ADJUST COLUMNS
###################################################################
def auto_adjust_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value) if cell.value else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = max_len + 2

###################################################################
# 7. INTERFAZ (TK) + IMAGEN DE FONDO
###################################################################
root = tk.Tk()
root.title("F5 + Cloudflare (Estilo Matrix) - Debug Excel")
window_width = 1200
window_height = 800
root.geometry(f"{window_width}x{window_height}")

def center_window(r):
    r.update_idletasks()
    screen_w = r.winfo_screenwidth()
    screen_h = r.winfo_screenheight()
    x = (screen_w - window_width) // 2
    y = (screen_h - window_height) // 2
    r.geometry(f"{window_width}x{window_height}+{x}+{y}")

center_window(root)
root.resizable(True, True)

try:
    bg_image = Image.open("fondo.jpg")  # Ajusta a tu imagen
    bg_image = bg_image.resize((window_width, window_height), Resampling.LANCZOS)
    bg_photo = ImageTk.PhotoImage(bg_image)
    bg_label = tk.Label(root, image=bg_photo)
    bg_label.place(x=0, y=0, relwidth=1, relheight=1)
except Exception as e:
    logging.warning(f"No se pudo cargar la imagen de fondo: {e}")

MATRIX_FONT = ("Courier", 14)
style = ttk.Style()
style.theme_use("clam")
root.configure(bg="black")
style.configure("Matrix.TNotebook", background="black")
style.configure("Matrix.TFrame", background="black")
style.configure("Matrix.TLabel", background="black", foreground="green", font=MATRIX_FONT)
style.configure("Matrix.TButton", background="black", foreground="green", font=MATRIX_FONT)
style.configure("Matrix.TCheckbutton", background="black", foreground="green", font=MATRIX_FONT)
style.configure("Matrix.TNotebook.Tab", background="black", foreground="green", font=MATRIX_FONT)

nb = ttk.Notebook(root, style="Matrix.TNotebook")
nb.pack(fill="both", expand=True)

# ------------- PESTAÑA LOGIN F5 -------------
frame_login_f5 = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_login_f5, text="LOGIN F5")
lbl_ip = ttk.Label(frame_login_f5, text="IP del F5:", style="Matrix.TLabel")
lbl_ip.pack(anchor="center", pady=5)
entry_login_ip = tk.Entry(frame_login_f5, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center")
entry_login_ip.pack(anchor="center", pady=5)
lbl_user_f5 = ttk.Label(frame_login_f5, text="Usuario F5:", style="Matrix.TLabel")
lbl_user_f5.pack(anchor="center", pady=5)
entry_login_user = tk.Entry(frame_login_f5, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center")
entry_login_user.pack(anchor="center", pady=5)
lbl_pass_f5 = ttk.Label(frame_login_f5, text="Contraseña F5:", style="Matrix.TLabel")
lbl_pass_f5.pack(anchor="center", pady=5)
entry_login_pass = tk.Entry(frame_login_f5, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center", show="*")
entry_login_pass.pack(anchor="center", pady=5)
btn_save_login_f5 = ttk.Button(frame_login_f5, text="Guardar Login F5", style="Matrix.TButton", command=save_login_f5)
btn_save_login_f5.pack(anchor="center", pady=10)

# ------------- PESTAÑA BASICO (F5) -------------
frame_basic = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_basic, text="BÁSICO (F5)")
lbl_partition = ttk.Label(frame_basic, text="Partition (opcional):", style="Matrix.TLabel")
lbl_partition.pack(anchor="center", pady=5)
entry_part_b = tk.Entry(frame_basic, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center")
entry_part_b.pack(anchor="center", pady=5)
btn_basic = ttk.Button(frame_basic, text="Obtener Datos (Básico)", style="Matrix.TButton", command=run_basic_collection)
btn_basic.pack(anchor="center", pady=10)

# ------------- PESTAÑA AVANZADA (F5) -------------
frame_adv = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_adv, text="AVANZADA (F5)")
lbl_adv = ttk.Label(frame_adv, text="(Usará la IP, Usuario y Contraseña guardados en LOGIN F5)", style="Matrix.TLabel")
lbl_adv.pack(anchor="center", pady=10)
btn_adv = ttk.Button(frame_adv, text="Obtener Info Avanzada", style="Matrix.TButton", command=run_advanced_collection)
btn_adv.pack(anchor="center", pady=10)

# ------------- PESTAÑA MONITORES (F5) -------------
frame_mon = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_mon, text="MONITORES (F5)")
btn_mon = ttk.Button(frame_mon, text="Obtener Datos de Monitores", style="Matrix.TButton", command=run_monitors_collection)
btn_mon.pack(anchor="center", pady=10)

# ------------- PESTAÑA LOGIN CLOUDFLARE -------------
frame_login_cf = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_login_cf, text="LOGIN Cloudflare")
lbl_cf_email = ttk.Label(frame_login_cf, text="Email Cloudflare:", style="Matrix.TLabel")
lbl_cf_email.pack(anchor="center", pady=5)
entry_cf_email = tk.Entry(frame_login_cf, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center")
entry_cf_email.pack(anchor="center", pady=5)
lbl_cf_token = ttk.Label(frame_login_cf, text="API Key/Token Cloudflare:", style="Matrix.TLabel")
lbl_cf_token.pack(anchor="center", pady=5)
entry_cf_token = tk.Entry(frame_login_cf, bg="black", fg="green", insertbackground="green", font=MATRIX_FONT, justify="center", show="*")
entry_cf_token.pack(anchor="center", pady=5)
btn_save_login_cf = ttk.Button(frame_login_cf, text="Guardar Login CF", style="Matrix.TButton", command=save_login_cf)
btn_save_login_cf.pack(anchor="center", pady=10)

# ------------- PESTAÑA CLOUDFLARE (consulta) -------------
frame_cf = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_cf, text="CLOUDFLARE")
lbl_cf_info = ttk.Label(frame_cf, text="(Usará Email y Token guardados en LOGIN CF)", style="Matrix.TLabel")
lbl_cf_info.pack(anchor="center", pady=10)
btn_cf = ttk.Button(frame_cf, text="Consultar Cloudflare", style="Matrix.TButton", command=run_cloudflare_collection)
btn_cf.pack(anchor="center", pady=10)

# ------------- PESTAÑA CF-EXTRA -------------
frame_cf_extra = ttk.Frame(nb, style="Matrix.TFrame", padding="10 10 10 10")
nb.add(frame_cf_extra, text="CF-EXTRA")
btn_cf_extra = ttk.Button(frame_cf_extra, text="Obtener Datos Extra", style="Matrix.TButton", command=run_cf_extra_collection)
btn_cf_extra.pack(anchor="center", pady=10)

# ------------- PESTAÑA RESULTADOS -------------
frame_results = ttk.Frame(nb, style="Matrix.TFrame", padding="5 5 5 5")
nb.add(frame_results, text="RESULTADOS")
frame_results.rowconfigure(0, weight=1)
frame_results.columnconfigure(0, weight=1)
text_area = scrolledtext.ScrolledText(frame_results, wrap=tk.WORD)
text_area.grid(row=0, column=0, sticky="nsew")
text_area.configure(bg="black", fg="green", insertbackground="green", font=MATRIX_FONT)

# ------------- FRAME INFERIOR (SSL & Exportar) -------------
frame_bottom = ttk.Frame(root, style="Matrix.TFrame", padding="5 5 5 5")
frame_bottom.pack(fill="x")
ssl_verify_var = tk.BooleanVar(value=True)
check_ssl = ttk.Checkbutton(frame_bottom, text="Verificar Certificado SSL", variable=ssl_verify_var, style="Matrix.TCheckbutton")
check_ssl.pack(side="left", padx=10)

def select_ca_file():
    global CA_FILE_PATH
    file_path = filedialog.askopenfilename(filetypes=[("Certificado CA", "*.crt *.pem"), ("Todos los archivos", "*.*")])
    if file_path:
        CA_FILE_PATH = file_path
        label_ca.config(text=f"CA: {file_path}")
    else:
        CA_FILE_PATH = None
        label_ca.config(text="CA: (No seleccionado)")

btn_ca = ttk.Button(frame_bottom, text="Seleccionar CA File", style="Matrix.TButton", command=select_ca_file)
btn_ca.pack(side="left", padx=10)
label_ca = ttk.Label(frame_bottom, text="CA: (No seleccionado)", style="Matrix.TLabel")
label_ca.pack(side="left", padx=10)

btn_export = ttk.Button(frame_bottom, text="Exportar a Excel", style="Matrix.TButton", command=export_to_excel)
btn_export.pack(side="right", padx=10)
progress_bar = ttk.Progressbar(frame_bottom, orient="horizontal", length=250, mode="determinate")
progress_bar.pack(side="right", padx=10)
banner_label = ttk.Label(root, text="Desarrollado por Alejandro Cuzmar", style="Matrix.TLabel")
banner_label.pack(side="bottom", pady=5)

root.mainloop()

